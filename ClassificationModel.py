# encoding: utf-8
# time    : 2021-06-02 15:09
# author  : 丛星星
"""
用于新闻文本分类。
"""
import tensorflow as tf
from tensorflow.python.platform import gfile
import numpy as np
import openpyxl
import bert_tokenization
from preprocess import move_stop_words

class ClassificationModel(object):
    def __init__(self):
        self.tokenizer = None
        self.sess = None
        self.is_train = None
        self.input_ids = None
        self.input_mask = None
        self.segment_ids = None
        self.predictions = None
        self.scores = None
        self.max_seq_length = None
        self.label_dict = ["财经", "房产", "教育", "科技", "体育", "游戏", "娱乐", "军事", "汽车", "其他"]  #

    # 加载训练模型
    def load_model(self, vocab_file, model_path, max_seq_length):
        self.tokenizer = bert_tokenization.FullTokenizer(vocab_file=vocab_file, do_lower_case=True)
        # 如果使用CPU
        sess_config = tf.ConfigProto(allow_soft_placement=True)
        self.sess = tf.Session(config=sess_config)
        with gfile.FastGFile(model_path, "rb") as f:
            graph_def = tf.GraphDef()
            graph_def.ParseFromString(f.read())
            self.sess.graph.as_default()
            tf.import_graph_def(graph_def, name="")

        self.sess.run(tf.global_variables_initializer())
        self.is_train = self.sess.graph.get_tensor_by_name("input/is_train:0")
        self.input_ids = self.sess.graph.get_tensor_by_name("input/input_ids:0")
        self.input_mask = self.sess.graph.get_tensor_by_name("input/input_mask:0")
        self.segment_ids = self.sess.graph.get_tensor_by_name("input/segment_ids:0")
        self.predictions = self.sess.graph.get_tensor_by_name("output_layer/predictions:0")
        self.scores = self.sess.graph.get_tensor_by_name("output_layer/scores:0")
        self.max_seq_length = max_seq_length

    # Tokenization处理
    def convert_feature_with_title(self, content, title):
        max_seq_length = self.max_seq_length
        title_token = self.tokenizer.tokenize(title)
        title_len = len(title_token)
        max_length_context = max_seq_length - 3 - title_len
        content_token = self.tokenizer.tokenize(content)
        if len(content_token) > max_length_context:
            content_token = content_token[:max_length_context]

        tokens = []
        segment_ids = []
        tokens.append("[CLS]")
        segment_ids.append(0)

        for token in title_token:
            tokens.append(token)
            segment_ids.append(0)

        tokens.append("[SEP]")
        segment_ids.append(0)

        for token in content_token:
            tokens.append(token)
            segment_ids.append(1)
        tokens.append("[SEP]")
        segment_ids.append(1)

        input_ids = self.tokenizer.convert_tokens_to_ids(tokens)
        input_mask = [1] * len(input_ids)
        while len(input_ids) < max_seq_length:
            input_ids.append(0)
            input_mask.append(0)
            segment_ids.append(1)
        assert len(input_ids) == max_seq_length
        assert len(input_mask) == max_seq_length
        assert len(segment_ids) == max_seq_length
        input_ids = np.array(input_ids)
        input_mask = np.array(input_mask)
        segment_ids = np.array(segment_ids)
        return input_ids, input_mask, segment_ids

    # 单条预测
    def predict_with_title(self, content, title):
        input_ids_temp, input_mask_temp, segment_ids_temp = self.convert_feature_with_title(content, title)
        feed = {self.is_train: False,
                self.input_ids: input_ids_temp.reshape(1, self.max_seq_length),
                self.input_mask: input_mask_temp.reshape(1, self.max_seq_length),
                self.segment_ids: segment_ids_temp.reshape(1, self.max_seq_length)}
        [label, score] = self.sess.run([self.predictions, self.scores], feed)
        label_name = self.label_dict[label[0]]
        return label_name, score

    # 批量预测
    def predict_label_write_to_file(self, file_path):
        # 读取excel文件对象
        read_workbook = openpyxl.load_workbook(file_path)
        read_worksheet = read_workbook.worksheets[0]
        # 写入excel文件对象
        write_workbook = openpyxl.Workbook()
        write_worksheet = write_workbook.active
        # 遍历数据
        rows = read_worksheet.iter_rows()
        index = 1
        for row in rows:
            # 新闻编号
            num = row[0].value
            # 新闻标签
            channel_name = 'channelName'
            # 新闻标题
            title = str(row[2].value)
            # 新闻内容
            content = str(row[3].value)
            # 第一行跳过
            if index == 1:
                write_worksheet.cell(index, 1, num)
                write_worksheet.cell(index, 2, channel_name)
                write_worksheet.cell(index, 3, title)
                write_worksheet.cell(index, 4, content)
                index += 1
                continue
            # 使用双输入预测方法
            predict_label_name, score = self.predict_with_title(content=move_stop_words(content), title=move_stop_words(title))
            if score[0][np.argmax(score)] < 0.8:
                predict_label_name = "其他"
            # 预测结果保存
            write_worksheet.cell(index, 1, num)
            write_worksheet.cell(index, 2, predict_label_name)
            write_worksheet.cell(index, 3, title)
            write_worksheet.cell(index, 4, content)
            index += 1
        # 保存预测标签
        write_workbook.save(file_path)


    def predict_label_with_logical(self, sentence):
        key_words = []
        pass
